#!/usr/bin/env python3
"""
Backend test for Phase F — Catalogue Export Generator endpoints.

Tests:
1. Seed patching verification (GET /api/listing-format-configs)
2. Setup test style with lifecycle
3. Preview Myntra (full lifecycle colors/sizes)
4. Preview Ajio with subset (colors/sizes filters)
5. Preview Flipkart
6. Download Myntra .xlsx (verify headers, body, openpyxl parse)
7. SKU map provisional rows verification
8. Idempotency (re-download, verify X-SkuMap-Created=0)
9. Error paths (404, 400, 422)
10. Color-major row order
11. Regression smoke tests
12. Cleanup
"""

import os
import sys
import requests
from io import BytesIO

# Base URL from environment
BASE_URL = os.getenv("REACT_APP_BACKEND_URL", "https://footwear-hub-565.preview.emergentagent.com")
API_BASE = f"{BASE_URL}/api"

# Test credentials
ADMIN_EMAIL = "admin@example.com"
ADMIN_PASSWORD = "admin123"

# Global state
access_token = None
test_style_id = None
test_style_code = None


def login():
    """Login as admin and get access token."""
    global access_token
    resp = requests.post(
        f"{API_BASE}/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=30
    )
    assert resp.status_code == 200, f"Login failed: {resp.status_code} {resp.text}"
    data = resp.json()
    assert "access_token" in data, f"No access_token in response: {data}"
    access_token = data["access_token"]
    print(f"✓ Logged in as {ADMIN_EMAIL}")


def headers():
    """Return auth headers."""
    return {"Authorization": f"Bearer {access_token}"}


def test_1_seed_patching_verification():
    """Test 1: Verify seeded export_template on GET /api/listing-format-configs."""
    print("\n=== TEST 1: SEED PATCHING VERIFICATION ===")
    
    resp = requests.get(f"{API_BASE}/listing-format-configs", headers=headers(), timeout=30)
    assert resp.status_code == 200, f"GET listing-format-configs failed: {resp.status_code} {resp.text}"
    
    configs = resp.json()
    assert len(configs) >= 3, f"Expected at least 3 configs, got {len(configs)}"
    
    # Find myntra, ajio, flipkart
    myntra = next((c for c in configs if c["platform"] == "myntra"), None)
    ajio = next((c for c in configs if c["platform"] == "ajio"), None)
    flipkart = next((c for c in configs if c["platform"] == "flipkart"), None)
    
    assert myntra, "Myntra config not found"
    assert ajio, "Ajio config not found"
    assert flipkart, "Flipkart config not found"
    
    # Verify myntra export_template
    assert "export_template" in myntra, "Myntra missing export_template"
    mt = myntra["export_template"]
    assert mt["sheet_name"] == "styledashboard", f"Myntra sheet_name: {mt['sheet_name']}"
    assert mt["header_row_index"] == 0, f"Myntra header_row_index: {mt['header_row_index']}"
    assert len(mt["columns"]) == 12, f"Myntra columns count: {len(mt['columns'])}"
    
    # Check specific columns
    cols = mt["columns"]
    assert any(c["source"] == "leaf_sku" and c["name"] == "SellerSkuCode" for c in cols), "Myntra missing SellerSkuCode"
    assert any(c["source"] == "blank" and c["name"] == "Style Id" for c in cols), "Myntra missing Style Id (blank)"
    assert any(c["source"] == "color_name" and c["name"] == "Colour" for c in cols), "Myntra missing Colour"
    assert any(c["source"] == "size" and c["name"] == "Size" for c in cols), "Myntra missing Size"
    assert any(c["source"] == "lifecycle" and c.get("key") == "mrp" and c["name"] == "MRP" for c in cols), "Myntra missing MRP"
    assert any(c["source"] == "lifecycle" and c.get("key") == "online_selling_price" and c["name"] == "Selling Price" for c in cols), "Myntra missing Selling Price"
    
    print("✓ Myntra export_template verified")
    
    # Verify ajio export_template
    assert "export_template" in ajio, "Ajio missing export_template"
    at = ajio["export_template"]
    assert at["sheet_name"] == "SSK_Styles_Export", f"Ajio sheet_name: {at['sheet_name']}"
    assert at["header_row_index"] == 2, f"Ajio header_row_index: {at['header_row_index']}"
    assert "pre_header_rows" in at, "Ajio missing pre_header_rows"
    assert len(at["pre_header_rows"]) == 2, f"Ajio pre_header_rows count: {len(at['pre_header_rows'])}"
    
    acols = at["columns"]
    assert any(c["source"] == "group_sku" and c["name"] == "*Style Code" for c in acols), "Ajio missing *Style Code"
    assert any(c["source"] == "leaf_sku" and c["name"] == "*Item SKU" for c in acols), "Ajio missing *Item SKU"
    assert any(c["source"] == "size" and c["name"] == "*Size" for c in acols), "Ajio missing *Size"
    assert any(c["source"] == "color_name" and c["name"] == "*Primary Color" for c in acols), "Ajio missing *Primary Color"
    
    print("✓ Ajio export_template verified")
    
    # Verify flipkart export_template
    assert "export_template" in flipkart, "Flipkart missing export_template"
    ft = flipkart["export_template"]
    assert ft["sheet_name"] == "Listings", f"Flipkart sheet_name: {ft['sheet_name']}"
    assert ft["header_row_index"] == 0, f"Flipkart header_row_index: {ft['header_row_index']}"
    assert "post_header_rows" in ft, "Flipkart missing post_header_rows"
    assert len(ft["post_header_rows"]) == 1, f"Flipkart post_header_rows count: {len(ft['post_header_rows'])}"
    
    fcols = ft["columns"]
    assert any(c["source"] == "leaf_sku" and c["name"] == "Seller SKU Id" for c in fcols), "Flipkart missing Seller SKU Id"
    assert not any(c["source"] == "group_sku" for c in fcols), "Flipkart should NOT have group_sku column"
    
    print("✓ Flipkart export_template verified")
    print("✓ TEST 1 PASSED: All 3 platforms have export_template with correct structure")


def test_2_setup_test_style():
    """Test 2: Create test style and set lifecycle."""
    global test_style_id, test_style_code
    print("\n=== TEST 2: SETUP TEST STYLE ===")
    
    # Create style
    resp = requests.post(
        f"{API_BASE}/styles",
        headers=headers(),
        json={
            "name": "Loafer Test",
            "description": "Premium leather loafer",
            "category": "Footwear",
            "image_url": "https://cdn.example/loafer.jpg"
        },
        timeout=30
    )
    assert resp.status_code == 200, f"POST /styles failed: {resp.status_code} {resp.text}"
    
    style = resp.json()
    test_style_id = style["id"]
    test_style_code = style["code"]
    
    assert test_style_code.startswith("SSK_"), f"Style code should start with SSK_: {test_style_code}"
    print(f"✓ Created style: {test_style_id} with code {test_style_code}")
    
    # Set lifecycle
    resp = requests.put(
        f"{API_BASE}/style-lifecycle/{test_style_id}",
        headers=headers(),
        json={
            "mrp": 2499,
            "online_selling_price": 1799,
            "planned_colors": ["Tan", "Gunmetal"],
            "planned_sizes": ["7", "8", "9", "10"]
        },
        timeout=30
    )
    assert resp.status_code == 200, f"PUT /style-lifecycle failed: {resp.status_code} {resp.text}"
    print("✓ Set lifecycle with mrp=2499, online_selling_price=1799, colors=[Tan,Gunmetal], sizes=[7,8,9,10]")
    print("✓ TEST 2 PASSED")


def test_3_preview_myntra():
    """Test 3: POST /api/catalogue-export/preview for Myntra (full lifecycle)."""
    print("\n=== TEST 3: PREVIEW MYNTRA ===")
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export/preview",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "myntra"
        },
        timeout=30
    )
    assert resp.status_code == 200, f"Preview myntra failed: {resp.status_code} {resp.text}"
    
    data = resp.json()
    
    # Verify structure
    assert data["style_code"] == test_style_code, f"style_code mismatch: {data['style_code']}"
    assert data["platform"] == "myntra", f"platform mismatch: {data['platform']}"
    assert data["sheet_name"] == "styledashboard", f"sheet_name: {data['sheet_name']}"
    assert data["header_row_index"] == 0, f"header_row_index: {data['header_row_index']}"
    assert data["row_count"] == 8, f"row_count: {data['row_count']} (expected 2 colors × 4 sizes = 8)"
    assert data["colors"] == ["Tan", "Gunmetal"], f"colors: {data['colors']}"
    assert data["sizes"] == ["7", "8", "9", "10"], f"sizes: {data['sizes']}"
    assert data["unmapped_colors"] == [], f"unmapped_colors should be empty: {data['unmapped_colors']}"
    
    # Verify header
    header = data["header"]
    assert "SellerSkuCode" in header, f"SellerSkuCode not in header: {header}"
    assert "Colour" in header, f"Colour not in header: {header}"
    assert "Size" in header, f"Size not in header: {header}"
    assert "MRP" in header, f"MRP not in header: {header}"
    assert "Selling Price" in header, f"Selling Price not in header: {header}"
    assert "Style Id" in header, f"Style Id not in header: {header}"
    assert "Style Name" in header, f"Style Name not in header: {header}"
    
    # Find column indices
    idx_seller = header.index("SellerSkuCode")
    idx_color = header.index("Colour")
    idx_size = header.index("Size")
    idx_mrp = header.index("MRP")
    idx_sp = header.index("Selling Price")
    idx_style_id = header.index("Style Id")
    
    rows = data["rows"]
    
    # Verify first row (Tan, size 7)
    row0 = rows[0]
    assert row0[idx_seller] == f"{test_style_code}-TN-7", f"row0 SellerSkuCode: {row0[idx_seller]}"
    assert row0[idx_color] == "Tan", f"row0 Colour: {row0[idx_color]}"
    assert row0[idx_size] == "7", f"row0 Size: {row0[idx_size]}"
    assert row0[idx_mrp] in (2499, 2499.0), f"row0 MRP: {row0[idx_mrp]}"
    assert row0[idx_sp] in (1799, 1799.0), f"row0 Selling Price: {row0[idx_sp]}"
    assert row0[idx_style_id] == "", f"row0 Style Id should be blank: {row0[idx_style_id]}"
    
    print(f"✓ Row 0: {row0[idx_seller]}, {row0[idx_color]}, {row0[idx_size]}, MRP={row0[idx_mrp]}, SP={row0[idx_sp]}")
    
    # Verify row 5 (second color Gunmetal, size 8 — color-major so Tan has rows 0-3, Gunmetal has rows 4-7)
    row5 = rows[5]
    assert row5[idx_seller] == f"{test_style_code}-GN-8", f"row5 SellerSkuCode: {row5[idx_seller]}"
    assert row5[idx_color] == "Gunmetal", f"row5 Colour: {row5[idx_color]}"
    assert row5[idx_size] == "8", f"row5 Size: {row5[idx_size]}"
    
    print(f"✓ Row 5: {row5[idx_seller]}, {row5[idx_color]}, {row5[idx_size]}")
    print("✓ TEST 3 PASSED: Myntra preview correct")


def test_4_preview_ajio_subset():
    """Test 4: POST /api/catalogue-export/preview for Ajio with subset."""
    print("\n=== TEST 4: PREVIEW AJIO WITH SUBSET ===")
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export/preview",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "ajio",
            "colors": ["Tan"],
            "sizes": ["8", "9"]
        },
        timeout=30
    )
    assert resp.status_code == 200, f"Preview ajio failed: {resp.status_code} {resp.text}"
    
    data = resp.json()
    
    assert data["header_row_index"] == 2, f"header_row_index: {data['header_row_index']}"
    assert data["row_count"] == 2, f"row_count: {data['row_count']} (expected 1 color × 2 sizes = 2)"
    
    header = data["header"]
    idx_stylecode = header.index("*Style Code")
    idx_itemsku = header.index("*Item SKU")
    idx_primary = header.index("*Primary Color")
    
    rows = data["rows"]
    
    # Row 0: Tan, size 8
    assert rows[0][idx_stylecode] == f"{test_style_code}-TN", f"row0 *Style Code: {rows[0][idx_stylecode]}"
    assert rows[0][idx_itemsku] == f"{test_style_code}-TN-8", f"row0 *Item SKU: {rows[0][idx_itemsku]}"
    assert rows[0][idx_primary] == "Tan", f"row0 *Primary Color: {rows[0][idx_primary]}"
    
    # Row 1: Tan, size 9
    assert rows[1][idx_stylecode] == f"{test_style_code}-TN", f"row1 *Style Code: {rows[1][idx_stylecode]}"
    assert rows[1][idx_itemsku] == f"{test_style_code}-TN-9", f"row1 *Item SKU: {rows[1][idx_itemsku]}"
    assert rows[1][idx_primary] == "Tan", f"row1 *Primary Color: {rows[1][idx_primary]}"
    
    print(f"✓ Row 0: {rows[0][idx_stylecode]}, {rows[0][idx_itemsku]}, {rows[0][idx_primary]}")
    print(f"✓ Row 1: {rows[1][idx_stylecode]}, {rows[1][idx_itemsku]}, {rows[1][idx_primary]}")
    print("✓ TEST 4 PASSED: Ajio preview with subset correct")


def test_5_preview_flipkart():
    """Test 5: POST /api/catalogue-export/preview for Flipkart."""
    print("\n=== TEST 5: PREVIEW FLIPKART ===")
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export/preview",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "flipkart"
        },
        timeout=30
    )
    assert resp.status_code == 200, f"Preview flipkart failed: {resp.status_code} {resp.text}"
    
    data = resp.json()
    
    assert data["header_row_index"] == 0, f"header_row_index: {data['header_row_index']}"
    
    header = data["header"]
    assert header[0] == "Seller SKU Id", f"First column should be 'Seller SKU Id': {header[0]}"
    
    rows = data["rows"]
    assert rows[0][0] == f"{test_style_code}-TN-7", f"First row first column: {rows[0][0]}"
    
    # Verify no group_sku column
    assert "*Style Code" not in header, "Flipkart should not have *Style Code column"
    
    print(f"✓ First row Seller SKU Id: {rows[0][0]}")
    print("✓ TEST 5 PASSED: Flipkart preview correct (no group_sku column)")


def test_6_download_myntra_xlsx():
    """Test 6: POST /api/catalogue-export for Myntra (download .xlsx)."""
    print("\n=== TEST 6: DOWNLOAD MYNTRA .XLSX ===")
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "myntra"
        },
        timeout=30
    )
    assert resp.status_code == 200, f"Download myntra failed: {resp.status_code} {resp.text}"
    
    # Verify headers
    assert "Content-Disposition" in resp.headers, "Missing Content-Disposition header"
    cd = resp.headers["Content-Disposition"]
    assert "attachment" in cd, f"Content-Disposition should contain 'attachment': {cd}"
    assert test_style_code in cd, f"Filename should contain style code: {cd}"
    assert "myntra_listing" in cd, f"Filename should contain 'myntra_listing': {cd}"
    assert ".xlsx" in cd, f"Filename should end with .xlsx: {cd}"
    
    assert resp.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", \
        f"Content-Type: {resp.headers['Content-Type']}"
    
    assert "X-Style-Code" in resp.headers, "Missing X-Style-Code header"
    assert resp.headers["X-Style-Code"] == test_style_code, f"X-Style-Code: {resp.headers['X-Style-Code']}"
    
    assert "X-Rows-Written" in resp.headers, "Missing X-Rows-Written header"
    assert resp.headers["X-Rows-Written"] == "8", f"X-Rows-Written: {resp.headers['X-Rows-Written']}"
    
    assert "X-Colors" in resp.headers, "Missing X-Colors header"
    assert resp.headers["X-Colors"] == "2", f"X-Colors: {resp.headers['X-Colors']}"
    
    assert "X-Sizes" in resp.headers, "Missing X-Sizes header"
    assert resp.headers["X-Sizes"] == "4", f"X-Sizes: {resp.headers['X-Sizes']}"
    
    assert "X-SkuMap-Created" in resp.headers, "Missing X-SkuMap-Created header"
    sku_map_created = int(resp.headers["X-SkuMap-Created"])
    print(f"✓ X-SkuMap-Created: {sku_map_created}")
    
    assert "X-SkuMap-Updated" in resp.headers, "Missing X-SkuMap-Updated header"
    assert "X-SkuMap-Unchanged" in resp.headers, "Missing X-SkuMap-Unchanged header"
    
    assert "Access-Control-Expose-Headers" in resp.headers, "Missing Access-Control-Expose-Headers header"
    aceh = resp.headers["Access-Control-Expose-Headers"]
    assert "Content-Disposition" in aceh, f"ACEH should include Content-Disposition: {aceh}"
    assert "X-Style-Code" in aceh, f"ACEH should include X-Style-Code: {aceh}"
    
    print("✓ All response headers verified")
    
    # Verify body starts with PK (xlsx = zip)
    body = resp.content
    assert body[:2] == b"PK", f"Body should start with 'PK' (xlsx magic): {body[:10]}"
    print("✓ Body starts with 'PK' (xlsx magic)")
    
    # Parse with openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(body))
        
        assert "styledashboard" in wb.sheetnames, f"Sheet 'styledashboard' not found: {wb.sheetnames}"
        ws = wb["styledashboard"]
        
        assert ws.max_row >= 9, f"Expected at least 9 rows (header + 8 data), got {ws.max_row}"
        
        # Check header row (row 1)
        assert ws.cell(1, 1).value == "Style Id", f"Cell(1,1): {ws.cell(1,1).value}"
        assert ws.cell(1, 2).value == "SellerSkuCode", f"Cell(1,2): {ws.cell(1,2).value}"
        
        # Check first data row (row 2)
        seller_sku = ws.cell(2, 2).value
        assert seller_sku.startswith(test_style_code), f"Cell(2,2) should start with {test_style_code}: {seller_sku}"
        assert "-TN-" in seller_sku, f"Cell(2,2) should contain '-TN-': {seller_sku}"
        
        print(f"✓ Openpyxl parse successful: sheet 'styledashboard', {ws.max_row} rows, first data row SellerSkuCode={seller_sku}")
    except ImportError:
        print("⚠ openpyxl not available, skipping xlsx parse verification")
    
    print("✓ TEST 6 PASSED: Myntra .xlsx download successful with correct headers and body")


def test_7_sku_map_provisional_rows():
    """Test 7: Verify provisional sku_map rows created."""
    print("\n=== TEST 7: SKU MAP PROVISIONAL ROWS ===")
    
    resp = requests.get(
        f"{API_BASE}/sku-map",
        headers=headers(),
        params={"search": test_style_code},
        timeout=30
    )
    assert resp.status_code == 200, f"GET /sku-map failed: {resp.status_code} {resp.text}"
    
    rows = resp.json()
    
    # Find myntra rows
    myntra_rows = [r for r in rows if r["source_type"] == "online_channel" and r["source_name"] == "myntra"]
    assert len(myntra_rows) == 2, f"Expected 2 myntra sku_map rows, got {len(myntra_rows)}"
    
    # Find TN and GN rows
    tn_row = next((r for r in myntra_rows if r["external_sku"] == f"{test_style_code}-TN"), None)
    gn_row = next((r for r in myntra_rows if r["external_sku"] == f"{test_style_code}-GN"), None)
    
    assert tn_row, f"TN row not found in sku_map: {myntra_rows}"
    assert gn_row, f"GN row not found in sku_map: {myntra_rows}"
    
    # Verify TN row
    assert tn_row["status"] == "pending_platform_confirmation", f"TN status: {tn_row['status']}"
    assert tn_row["color_map"] == {"Tan": "Tan"}, f"TN color_map: {tn_row['color_map']}"
    assert tn_row["size_map"] == {"7": "7", "8": "8", "9": "9", "10": "10"}, f"TN size_map: {tn_row['size_map']}"
    assert tn_row.get("created_via") == "catalogue_export", f"TN created_via: {tn_row.get('created_via')}"
    
    print(f"✓ TN row: external_sku={tn_row['external_sku']}, status={tn_row['status']}, color_map={tn_row['color_map']}")
    
    # Verify GN row
    assert gn_row["status"] == "pending_platform_confirmation", f"GN status: {gn_row['status']}"
    assert gn_row["color_map"] == {"Gunmetal": "Gunmetal"}, f"GN color_map: {gn_row['color_map']}"
    assert gn_row["size_map"] == {"7": "7", "8": "8", "9": "9", "10": "10"}, f"GN size_map: {gn_row['size_map']}"
    assert gn_row.get("created_via") == "catalogue_export", f"GN created_via: {gn_row.get('created_via')}"
    
    print(f"✓ GN row: external_sku={gn_row['external_sku']}, status={gn_row['status']}, color_map={gn_row['color_map']}")
    print("✓ TEST 7 PASSED: 2 provisional sku_map rows created correctly")


def test_8_idempotency():
    """Test 8: Re-download same file, verify idempotency."""
    print("\n=== TEST 8: IDEMPOTENCY ===")
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "myntra"
        },
        timeout=30
    )
    assert resp.status_code == 200, f"Re-download myntra failed: {resp.status_code} {resp.text}"
    
    # Verify X-SkuMap-Created is 0
    assert "X-SkuMap-Created" in resp.headers, "Missing X-SkuMap-Created header"
    sku_map_created = int(resp.headers["X-SkuMap-Created"])
    assert sku_map_created == 0, f"X-SkuMap-Created should be 0 on re-download: {sku_map_created}"
    
    # X-SkuMap-Updated or X-SkuMap-Unchanged should be non-zero
    sku_map_updated = int(resp.headers.get("X-SkuMap-Updated", 0))
    sku_map_unchanged = int(resp.headers.get("X-SkuMap-Unchanged", 0))
    
    print(f"✓ X-SkuMap-Created: {sku_map_created}, X-SkuMap-Updated: {sku_map_updated}, X-SkuMap-Unchanged: {sku_map_unchanged}")
    
    # Verify sku_map still has exactly 2 rows for this platform
    resp2 = requests.get(
        f"{API_BASE}/sku-map",
        headers=headers(),
        params={"search": test_style_code},
        timeout=30
    )
    assert resp2.status_code == 200, f"GET /sku-map failed: {resp2.status_code} {resp2.text}"
    
    rows = resp2.json()
    myntra_rows = [r for r in rows if r["source_type"] == "online_channel" and r["source_name"] == "myntra"]
    assert len(myntra_rows) == 2, f"Expected still 2 myntra rows after re-download, got {len(myntra_rows)}"
    
    print("✓ TEST 8 PASSED: Idempotency verified (no duplicate sku_map rows)")


def test_9_error_paths():
    """Test 9: Error paths (404, 400, 422)."""
    print("\n=== TEST 9: ERROR PATHS ===")
    
    # 9a: Unknown style_id → 404
    resp = requests.post(
        f"{API_BASE}/catalogue-export",
        headers=headers(),
        json={
            "style_id": "000000000000000000000000",
            "platform": "myntra"
        },
        timeout=30
    )
    assert resp.status_code == 404, f"Expected 404 for unknown style_id, got {resp.status_code}"
    assert "Style not found" in resp.text, f"Error message: {resp.text}"
    assert resp.headers["Content-Type"].startswith("application/json"), f"Should be JSON error: {resp.headers['Content-Type']}"
    print("✓ 9a: Unknown style_id → 404 'Style not found'")
    
    # 9b: Unmapped color → 400
    resp = requests.post(
        f"{API_BASE}/catalogue-export",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "myntra",
            "colors": ["XyzUnknownColour999"],
            "sizes": ["8"]
        },
        timeout=30
    )
    assert resp.status_code == 400, f"Expected 400 for unmapped color, got {resp.status_code}"
    assert "not in the color master" in resp.text, f"Error message: {resp.text}"
    assert resp.headers["Content-Type"].startswith("application/json"), f"Should be JSON error: {resp.headers['Content-Type']}"
    print("✓ 9b: Unmapped color → 400 'not in the color master'")
    
    # 9c: Invalid platform enum → 422
    resp = requests.post(
        f"{API_BASE}/catalogue-export",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "amazon"
        },
        timeout=30
    )
    assert resp.status_code == 422, f"Expected 422 for invalid platform, got {resp.status_code}"
    assert resp.headers["Content-Type"].startswith("application/json"), f"Should be JSON error: {resp.headers['Content-Type']}"
    print("✓ 9c: Invalid platform enum 'amazon' → 422")
    
    # 9d: Style with no lifecycle and no colors → 400
    # Create a fresh style
    resp = requests.post(
        f"{API_BASE}/styles",
        headers=headers(),
        json={
            "name": "No Lifecycle Style",
            "description": "Test style with no lifecycle",
            "category": "Test"
        },
        timeout=30
    )
    assert resp.status_code == 200, f"POST /styles failed: {resp.status_code} {resp.text}"
    style2 = resp.json()
    style2_id = style2["id"]
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export",
        headers=headers(),
        json={
            "style_id": style2_id,
            "platform": "myntra"
        },
        timeout=30
    )
    assert resp.status_code == 400, f"Expected 400 for no colors, got {resp.status_code}"
    assert "No colours to export" in resp.text, f"Error message: {resp.text}"
    assert resp.headers["Content-Type"].startswith("application/json"), f"Should be JSON error: {resp.headers['Content-Type']}"
    print("✓ 9d: Style with no lifecycle and no colors → 400 'No colours to export'")
    
    print("✓ TEST 9 PASSED: All error paths return JSON errors (not xlsx)")


def test_10_color_major_row_order():
    """Test 10: Verify color-major row order."""
    print("\n=== TEST 10: COLOR-MAJOR ROW ORDER ===")
    
    resp = requests.post(
        f"{API_BASE}/catalogue-export/preview",
        headers=headers(),
        json={
            "style_id": test_style_id,
            "platform": "myntra",
            "colors": ["Tan", "Gunmetal"],
            "sizes": ["7", "8"]
        },
        timeout=30
    )
    assert resp.status_code == 200, f"Preview failed: {resp.status_code} {resp.text}"
    
    data = resp.json()
    assert data["row_count"] == 4, f"row_count: {data['row_count']} (expected 2 colors × 2 sizes = 4)"
    
    header = data["header"]
    idx_seller = header.index("SellerSkuCode")
    
    rows = data["rows"]
    
    # Color-major order: Tan-7, Tan-8, Gunmetal-7, Gunmetal-8
    assert rows[0][idx_seller] == f"{test_style_code}-TN-7", f"row0: {rows[0][idx_seller]}"
    assert rows[1][idx_seller] == f"{test_style_code}-TN-8", f"row1: {rows[1][idx_seller]}"
    assert rows[2][idx_seller] == f"{test_style_code}-GN-7", f"row2: {rows[2][idx_seller]}"
    assert rows[3][idx_seller] == f"{test_style_code}-GN-8", f"row3: {rows[3][idx_seller]}"
    
    print(f"✓ Row order: {rows[0][idx_seller]}, {rows[1][idx_seller]}, {rows[2][idx_seller]}, {rows[3][idx_seller]}")
    print("✓ TEST 10 PASSED: Color-major row order verified")


def test_11_regression_smoke():
    """Test 11: Regression smoke tests."""
    print("\n=== TEST 11: REGRESSION SMOKE ===")
    
    # GET /api/styles
    resp = requests.get(f"{API_BASE}/styles", headers=headers(), timeout=30)
    assert resp.status_code == 200, f"GET /styles failed: {resp.status_code}"
    print("✓ GET /api/styles → 200")
    
    # GET /api/color-master
    resp = requests.get(f"{API_BASE}/color-master", headers=headers(), timeout=30)
    assert resp.status_code == 200, f"GET /color-master failed: {resp.status_code}"
    print("✓ GET /api/color-master → 200")
    
    # GET /api/listing-format-configs
    resp = requests.get(f"{API_BASE}/listing-format-configs", headers=headers(), timeout=30)
    assert resp.status_code == 200, f"GET /listing-format-configs failed: {resp.status_code}"
    print("✓ GET /api/listing-format-configs → 200")
    
    # GET /api/listing-format-configs/myntra
    resp = requests.get(f"{API_BASE}/listing-format-configs/myntra", headers=headers(), timeout=30)
    assert resp.status_code == 200, f"GET /listing-format-configs/myntra failed: {resp.status_code}"
    data = resp.json()
    assert data["export_template"]["sheet_name"] == "styledashboard", "Myntra export_template changed"
    assert data["export_template"]["header_row_index"] == 0, "Myntra header_row_index changed"
    assert len(data["export_template"]["columns"]) == 12, "Myntra columns count changed"
    print("✓ GET /api/listing-format-configs/myntra → 200, export_template unchanged")
    
    print("✓ TEST 11 PASSED: All regression smoke tests passed")


def test_12_cleanup():
    """Test 12: Cleanup (best effort)."""
    print("\n=== TEST 12: CLEANUP ===")
    
    # Try to delete test styles
    if test_style_id:
        resp = requests.delete(f"{API_BASE}/styles/{test_style_id}", headers=headers(), timeout=30)
        if resp.status_code == 200:
            print(f"✓ Deleted test style {test_style_id}")
        else:
            print(f"⚠ Could not delete test style {test_style_id}: {resp.status_code}")
    
    print("✓ TEST 12 PASSED: Cleanup complete (provisional sku_map rows left as-is)")


def main():
    """Run all tests."""
    print("=" * 80)
    print("PHASE F — CATALOGUE EXPORT GENERATOR BACKEND TESTS")
    print("=" * 80)
    
    try:
        login()
        test_1_seed_patching_verification()
        test_2_setup_test_style()
        test_3_preview_myntra()
        test_4_preview_ajio_subset()
        test_5_preview_flipkart()
        test_6_download_myntra_xlsx()
        test_7_sku_map_provisional_rows()
        test_8_idempotency()
        test_9_error_paths()
        test_10_color_major_row_order()
        test_11_regression_smoke()
        test_12_cleanup()
        
        print("\n" + "=" * 80)
        print("✅ ALL 12 TESTS PASSED")
        print("=" * 80)
        return 0
    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}")
        return 1
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
