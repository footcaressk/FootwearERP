"""Free / offline PO extraction using pdfplumber + openpyxl.

This module is the *primary* extractor — it never calls an LLM service, never
needs an internet round-trip, and never expires. If it cannot extract a usable
PO it raises ``ExtractionFailed`` so the caller can decide whether to fall back
to an LLM (the LLM path is kept in ``po_extractor.py`` as an opt-in backup).

The strategy is deterministic:
  1.  PDF -> pdfplumber: pull text + tables.  Excel -> openpyxl cells.
  2.  Header/meta fields are detected via labelled regex (PO No, Date, GST etc).
  3.  Line-item table is detected by finding header tokens such as
      "Article", "Style", "Size", "Quantity", "Rate", "Amount". Whichever row
      contains the most of these is treated as the column header.
  4.  Each subsequent row that has a numeric quantity & price becomes a line
      item. Size variants on the same row are flattened.
  5.  Tax + grand-total are inferred from the bottom block (CGST/SGST/IGST).
"""
import io
import re
from datetime import datetime
from typing import Optional

import openpyxl
import pdfplumber


class ExtractionFailed(Exception):
    pass


# ---------- common helpers ----------
_HSN_CODES_FOOTWEAR = "64029990"


def _to_number(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₹", "").replace("Rs", "").replace("INR", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else 0.0


def _to_int(v) -> int:
    return int(round(_to_number(v)))


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _normalise_date(s: str) -> str:
    """Convert DD.MM.YYYY / DD/MM/YYYY / DD-MM-YYYY to YYYY-MM-DD. Returns empty on failure."""
    if not s:
        return ""
    s = str(s).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Try regex: any 1-2 digits / 1-2 digits / 2-4 digits
    m = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        try:
            return datetime(int(y), int(mo), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _find_first(pattern: str, text: str, group: int = 1, flags=re.I) -> str:
    m = re.search(pattern, text, flags=flags)
    return _norm(m.group(group)) if m else ""


# ---------- PDF extraction ----------
def extract_po_from_pdf_local(file_bytes: bytes) -> dict:
    try:
        return _extract_pdf(file_bytes)
    except ExtractionFailed:
        raise
    except Exception as e:
        raise ExtractionFailed(f"PDF parse error: {e}") from e


def _extract_pdf(file_bytes: bytes) -> dict:
    full_text_parts = []
    all_tables = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            full_text_parts.append(t)
            for tab in (page.extract_tables() or []):
                if tab:
                    all_tables.append(tab)

    full_text = "\n".join(full_text_parts)
    if not full_text.strip():
        raise ExtractionFailed("PDF has no extractable text (scanned image?). Try Excel or LLM extractor.")

    data = _parse_meta(full_text)
    line_items = _parse_line_items_from_tables(all_tables, data.get("po_number", ""))
    # If the table parser found nothing meaningful, try a text-based fallback
    if not line_items:
        line_items = _parse_line_items_from_text(full_text)
    data["line_items"] = line_items

    # Totals: prefer explicit grand-total found in text; else compute
    _finalise_totals(data, full_text)
    return data


def _parse_meta(text: str) -> dict:
    """Return all header / metadata fields from free text."""
    # PO Number — try the strictest form first (digits-only or alphanum after explicit label)
    po_no = ""
    for pat in [
        r"(?:P\.?\s*O\.?|Purchase\s*Order)\s*(?:No\.?|#|Number)[\s:\-|]+([A-Z0-9][A-Z0-9\-_/]{3,})",
        r"\bOrder\s*(?:No\.?|#|Number)[\s:\-|]+([A-Z0-9][A-Z0-9\-_/]{3,})",
        r"\bP\.?O\.?\s*#[\s:\-|]*([A-Z0-9][A-Z0-9\-_/]{3,})",
    ]:
        m = re.search(pat, text, flags=re.I)
        if m:
            po_no = _norm(m.group(1))
            break

    po_date = _normalise_date(_find_first(r"(?:PO\s*Date|Order\s*Date|Date)[\s:\-|]+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", text))
    delivery_date = _normalise_date(_find_first(r"(?:Delivery|Ship(?:ment)?|Due)\s*Date[\s:\-|]+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", text))

    # Client / Buyer detection (tries in order):
    #   1. Explicit "Bill To" / "Buyer" / "Customer" label
    #   2. First non-empty line above the "PURCHASE ORDER" title (corporate header)
    client_name = ""
    for pat in [
        r"\b(?:Bill\s*To|Buyer|Customer|Consignee|Destination)(?:\s*Name)?\s*[\s:\-|]+\s*([A-Z][A-Za-z0-9 .,&'\-]{3,80})",
    ]:
        m = re.search(pat, text, flags=re.I)
        if m:
            client_name = _norm(m.group(1))
            # Filter out garbage like "VijayakanthJ ForSHEIN..."  (no spaces, camel-case run-ons)
            if " " in client_name and not re.search(r"\bSELLER\b|\bSignature\b|For[A-Z]{3,}", client_name, re.I):
                break
            client_name = ""

    if not client_name:
        # Look for the first ALL-CAPS line before "PURCHASE ORDER"
        lines = [_norm(ln) for ln in text.splitlines()]
        po_title_idx = next((i for i, ln in enumerate(lines) if re.search(r"PURCHASE\s*ORDER", ln, re.I)), -1)
        if po_title_idx > 0:
            for i in range(po_title_idx - 1, -1, -1):
                ln = lines[i]
                # Heuristic: a corporate name is upper-case-heavy, 4+ chars, contains a space or known suffix
                if (re.match(r"^[A-Z][A-Z0-9 &.,'\-/]{4,80}$", ln)
                        and any(k in ln.upper() for k in ["LIMITED", "LIMITED.", "LTD", "LLP", "PVT", "PRIVATE", "INC", "CO.", "CORP", "COMPANY"])):
                    client_name = ln
                    break
                if i == 0 and ln and ln.isupper():
                    client_name = ln
                    break

    # Vendor / seller detection — strongest signal first: line after "Vendor Code : <digits>"
    vendor_name = ""
    m = re.search(r"Vendor\s*Code\s*[:\-|]\s*\d+[^\n]*\n([A-Z][A-Z0-9 &.,'\-/]{4,80})", text, flags=re.I)
    if m:
        vendor_name = _norm(m.group(1))

    if not vendor_name:
        for pat in [
            r"\b(?:Vendor|Supplier|Sold\s*By)(?:\s*Name)?\s*[\s:\-|]+\s*([A-Z][A-Za-z0-9 .,&'\-]{3,80})",
            r"\bSeller(?:\s*Name)?\s*[\s:\-|]+\s*([A-Z][A-Za-z0-9 .,&'\-]{3,80})",
        ]:
            m = re.search(pat, text, flags=re.I)
            if m:
                cand = _norm(m.group(1))
                # Reject template/cover words
                if (" " in cand
                        and not re.search(r"\b(DRAFT|PURCHASE|ORDER|CODE|ACCEPT|SIGN|TITLE|DATE)\b", cand, re.I)):
                    vendor_name = cand
                    break

    return {
        "po_number": po_no or "",
        "po_date": po_date or "",
        "delivery_date": delivery_date or "",
        "client_name": client_name or "",
        "vendor_name": vendor_name or "",
        "client_address": "",
        "vendor_address": "",
        "billing_address": "",
        "shipping_address": "",
        "payment_terms": _find_first(r"Payment\s*Terms?\s*[:\-|]?\s*([^\n|]+)", text),
        "currency": "INR",
        "notes": "",
    }


# ---------- line-item table detection ----------
_HEADER_TOKENS = {
    "style": ["style", "article", "articleno", "article no", "model", "item code", "sku"],
    "description": ["description", "particulars", "item name", "product", "materialdescription", "material description"],
    "color": ["color", "colour", "shade"],
    "size": ["size", "uk size"],
    "hsn": ["hsn", "sac", "hsn code", "hsncode"],
    "quantity": ["quantity", "qty", "pcs", "pairs"],
    "unit_price": ["basecost", "base cost", "rate", "unit price", "unit rate"],
    "mrp": ["mrp"],
    "amount": ["totalbasevalue", "total base value", "amount", "total value", "net amount", "totalvalue"],
}

# Junk tokens we never want as a field
_HEADER_JUNK = {"sr.no", "sr no", "srno", "sr", "no", "uom", "ean", "ean no", "eanno",
                "vendoritemno", "vendor item no", "vendor article no", "vendorarticleno",
                "site", "deliverydate", "delivery date", "cess", "cess(%)", "cessfxdrt",
                "cessfxdvl", "igst", "igst(%)", "cgst", "sgst", "cgst(%)", "sgst(%)"}


def _classify_header(cell: str) -> str | None:
    s = _norm(cell).lower()
    if not s or s in _HEADER_JUNK:
        return None
    for key, candidates in _HEADER_TOKENS.items():
        for c in candidates:
            if c == s or (len(s) <= 22 and c in s):
                return key
    return None


def _split_color_size_from_desc(desc: str) -> tuple[str, str, str]:
    """If `desc` has comma-separated trailing pieces, take the last as size,
    second-to-last as color, the rest as description.

    Returns (description, color, size). Never raises."""
    if not desc:
        return "", "", ""
    parts = [p.strip() for p in desc.split(",")]
    if len(parts) >= 3:
        return ",".join(parts[:-2]).strip(), parts[-2], parts[-1]
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return desc, "", ""


def _parse_line_items_from_tables(tables: list, po_no: str) -> list[dict]:
    """Detect header row in each table. Handles multi-line cells where one
    visual cell contains multiple stacked sub-headers (eg. SHEIN's
    ``ArticleNo./HSNCode`` cell). For data rows, the corresponding sub-line is
    extracted from the same column."""
    items: list[dict] = []
    for tbl in tables:
        # Identify header row by counting classified sub-tokens in the first ~5 rows
        best_idx = -1
        best_score = 0
        # (col_idx, line_idx) -> field
        best_map: dict[tuple[int, int], str] = {}
        for i, row in enumerate(tbl[:6]):
            sub_classes: dict[tuple[int, int], str] = {}
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                for li, sub in enumerate(str(cell).splitlines()):
                    cls = _classify_header(sub)
                    if cls and (j, li) not in sub_classes:
                        sub_classes[(j, li)] = cls
            score = len(sub_classes)
            if score > best_score:
                best_score, best_idx, best_map = score, i, sub_classes
        if best_score < 2 or best_idx < 0:
            continue

        # Iterate data rows
        for row in tbl[best_idx + 1:]:
            if not row or all((c is None or str(c).strip() == "") for c in row):
                continue
            # Skip subtotal / footer-style rows
            first_cell = _norm(row[0]) if len(row) and row[0] is not None else ""
            if re.search(r"sub\s*total|grand\s*total|total\s*[a-z]", first_cell, re.I):
                continue

            rec = {"style_code": "", "description": "", "color": "", "size": "",
                   "hsn_code": "", "quantity": 0, "unit_price": 0.0,
                   "amount": 0.0, "mrp": ""}

            for (col_idx, line_idx), key in best_map.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None:
                    continue
                # Pick the corresponding sub-line of the data cell
                lines = str(val).splitlines()
                if line_idx < len(lines):
                    sval = _norm(lines[line_idx])
                else:
                    sval = _norm(val)
                if not sval:
                    continue
                if key == "style":
                    rec["style_code"] = sval
                elif key == "description":
                    rec["description"] = sval
                elif key == "color":
                    rec["color"] = sval
                elif key == "size":
                    rec["size"] = sval
                elif key == "hsn":
                    rec["hsn_code"] = sval
                elif key == "quantity":
                    rec["quantity"] = _to_int(sval)
                elif key == "unit_price":
                    rec["unit_price"] = _to_number(sval)
                elif key == "mrp":
                    rec["mrp"] = sval
                elif key == "amount":
                    rec["amount"] = _to_number(sval)

            # If description holds color+size embedded (eg. "SHEINWOMEN...,BLACK,3")
            # and color/size are still empty, split them out.
            if rec["description"] and not (rec["color"] and rec["size"]):
                desc, color, size = _split_color_size_from_desc(rec["description"])
                if size and not rec["size"]:
                    rec["size"] = size
                if color and not rec["color"]:
                    rec["color"] = color
                rec["description"] = desc or rec["description"]

            # Accept the row only if it has a positive qty + price + identifier
            if rec["quantity"] > 0 and rec["unit_price"] > 0 and (rec["style_code"] or rec["description"]):
                if not rec["hsn_code"]:
                    rec["hsn_code"] = _HSN_CODES_FOOTWEAR
                if not rec["amount"]:
                    rec["amount"] = round(rec["quantity"] * rec["unit_price"], 2)
                items.append(rec)
    return items


# ---------- text fallback for line items ----------
def _parse_line_items_from_text(text: str) -> list[dict]:
    """Lightweight fallback: look for repeating rows like 'STYLECODE  DESC  COLOR  SIZE  QTY  PRICE  AMOUNT'."""
    items = []
    for raw in text.splitlines():
        line = _norm(raw)
        # Heuristic: at least 4 tokens, with 2 large numbers near the end
        m = re.match(
            r"^([A-Z][A-Z0-9_\-]{2,})\s+(.+?)\s+(\d+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$",
            line,
        )
        if m:
            code, desc, qty, price, amount = m.groups()
            items.append({
                "style_code": code, "description": desc, "color": "", "size": "",
                "hsn_code": _HSN_CODES_FOOTWEAR,
                "quantity": int(qty),
                "unit_price": _to_number(price), "amount": _to_number(amount), "mrp": "",
            })
    return items


def _finalise_totals(data: dict, full_text: str) -> None:
    items = data.get("line_items", [])
    subtotal = sum(li.get("amount", 0) for li in items)
    total_qty = sum(li.get("quantity", 0) for li in items)

    # Try to detect explicit tax info
    cgst_amt = _to_number(_find_first(r"CGST[^\d-]*([0-9.,]+)", full_text))
    sgst_amt = _to_number(_find_first(r"SGST[^\d-]*([0-9.,]+)", full_text))
    igst_amt = _to_number(_find_first(r"\bTOTAL\s*IGST[\s:\-|]+(?:INR|Rs\.?)?[\s]*([0-9][0-9.,]+)", full_text))
    if not igst_amt:
        igst_amt = _to_number(_find_first(r"\bIGST[^\d-]*([0-9.,]+)", full_text))

    # Grand-total — try several variants in order of specificity
    grand = 0.0
    for pat in [
        r"Total\s*Order\s*Value\s*[:\-]?\s*(?:INR|Rs\.?)?\s*([0-9][0-9,]+(?:\.\d+)?)",
        r"TOTAL\s*BASIC\s*VALUE\s*(?:INR|Rs\.?)?\s*([0-9][0-9,]+(?:\.\d+)?)",
        r"(?:Grand\s*Total|Net\s*Payable|Total\s*Amount|Net\s*Value)\s*[:\-]?\s*(?:INR|Rs\.?)?\s*([0-9][0-9,]+(?:\.\d+)?)",
    ]:
        m = re.search(pat, full_text, flags=re.I)
        if m:
            grand = _to_number(m.group(1))
            break

    data["subtotal"] = round(subtotal, 2)
    data["total_quantity"] = total_qty
    data["cgst_rate"] = 0
    data["sgst_rate"] = 0
    data["igst_rate"] = 0
    data["cgst_amount"] = round(cgst_amt, 2)
    data["sgst_amount"] = round(sgst_amt, 2)
    data["igst_amount"] = round(igst_amt, 2)
    data["total_tax"] = round(cgst_amt + sgst_amt + igst_amt, 2)
    data["grand_total"] = round(grand if grand else subtotal + cgst_amt + sgst_amt + igst_amt, 2)


# ---------- Excel extraction ----------
def extract_po_from_xlsx_local(file_bytes: bytes) -> dict:
    try:
        return _extract_xlsx(file_bytes)
    except ExtractionFailed:
        raise
    except Exception as e:
        raise ExtractionFailed(f"Excel parse error: {e}") from e


def _extract_xlsx(file_bytes: bytes) -> dict:
    bio = io.BytesIO(file_bytes)
    try:
        wb = openpyxl.load_workbook(bio, data_only=True)
    except Exception as e:
        raise ExtractionFailed(f"Cannot open xlsx: {e}") from e

    # We build a (row, col) -> value map for the first sheet (and merge a flat textual rep for regex)
    ws = wb[wb.sheetnames[0]]
    grid = []
    flat_lines = []
    for row in ws.iter_rows(values_only=True):
        grid.append([("" if c is None else c) for c in row])
        flat_lines.append(" | ".join(("" if c is None else str(c)) for c in row))
    flat = "\n".join(flat_lines)

    data = _parse_meta(flat)
    items = _parse_xlsx_line_items(grid)
    if not items:
        items = _parse_line_items_from_text(flat)
    data["line_items"] = items
    _finalise_totals(data, flat)
    return data


def _parse_xlsx_line_items(grid: list[list]) -> list[dict]:
    """Detect the header row in the sheet and read rows below."""
    if not grid:
        return []
    best_row = -1
    best_map: dict[int, str] = {}
    best_score = 0
    # Header expected within first 30 rows
    for r_idx in range(min(len(grid), 30)):
        row = grid[r_idx]
        classes = [_classify_header(c) for c in row]
        score = sum(1 for c in classes if c)
        if score > best_score:
            best_score, best_row = score, r_idx
            best_map = {j: c for j, c in enumerate(classes) if c}

    if best_score < 2:
        return []

    items = []
    for r in grid[best_row + 1:]:
        if not any(str(c).strip() for c in r if c is not None):
            continue
        rec = {"style_code": "", "description": "", "color": "", "size": "",
               "hsn_code": "", "quantity": 0, "unit_price": 0.0, "amount": 0.0, "mrp": ""}
        for j, key in best_map.items():
            if j >= len(r):
                continue
            val = r[j]
            if key == "style":
                rec["style_code"] = _norm(val)
            elif key == "description":
                rec["description"] = _norm(val)
            elif key == "color":
                rec["color"] = _norm(val)
            elif key == "size":
                rec["size"] = _norm(val)
            elif key == "hsn":
                rec["hsn_code"] = _norm(val)
            elif key == "quantity":
                rec["quantity"] = _to_int(val)
            elif key == "unit_price":
                rec["unit_price"] = _to_number(val)
            elif key == "amount":
                rec["amount"] = _to_number(val)
        if rec["quantity"] > 0 and rec["unit_price"] > 0 and (rec["style_code"] or rec["description"]):
            if not rec["hsn_code"]:
                rec["hsn_code"] = _HSN_CODES_FOOTWEAR
            if not rec["amount"]:
                rec["amount"] = round(rec["quantity"] * rec["unit_price"], 2)
            items.append(rec)
    return items
