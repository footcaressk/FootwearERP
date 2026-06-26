"""Packing-list Excel generator.

Two modes:

1.  **Default SSK template** – mirrors the format the user uploaded (`Packing list SSK26-27-001.xlsx`).
    Vendor + destination metadata, PO header, carton dim, line items, grand-total,
    order-summary recap.

2.  **Per-client custom template** – an admin uploads any `.xlsx` and tags it
    with a client name. The file is stored verbatim with placeholder markers that
    will be substituted at generation time. Supported markers:

        {{po_number}}           PO number
        {{po_date}}             PO date (DD.MM.YYYY)
        {{client_name}}         Client / buyer name
        {{client_address}}      Client address
        {{vendor_name}}         Vendor / our company
        {{vendor_address}}      Vendor address
        {{vendor_gstin}}        Vendor GSTIN
        {{client_gstin}}        Buyer GSTIN
        {{carton_dim}}          Carton dimension (free text, persisted per PO)
        {{total_pcs}}           Total pieces
        {{total_cartons}}       Total cartons
        {{date}}                Today's date
        {{lines}}               Cell marker — the row containing this token is
                                 the first **line-item** row; the row layout is
                                 cloned downwards for each shipment line.

        Inside the {{lines}} row the columns are inferred by the *header row*
        immediately above it (e.g. cells with headers `Style`, `Colour`,
        `Size`, `Qty`, `Net Wt`, `Gross Wt`, sizes `36`-`42`, etc).
"""
from __future__ import annotations
import io
import re
from copy import copy
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Default SSK template -------------------------------------------------------
# ---------------------------------------------------------------------------
_DARK = "0F172A"
_ACCENT = "C27842"
_LIGHT = "F1F5F9"

_thin = Side(style="thin", color="64748B")
_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

VENDOR = {
    "name": "SSK FOOTCARE MANUFACTURING LLP",
    "address": "H 43, NARAYAN NIWAS, OPP JETVAN GARDEN, CHEMBUR MUMBAI, MAHARASHTRA - 400071",
    "gstin": "27AFKFS4410F1Z2",
}


def _set(ws, coord: str, val, *, bold=False, fill=None, color=None, size=10, align="left", border=False):
    cell = ws[coord]
    cell.value = val
    font_kwargs = {"name": "Helvetica", "size": size, "bold": bold}
    if color:
        font_kwargs["color"] = color
    cell.font = Font(**font_kwargs)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = _box


def _grid_border(ws, rng: str):
    for row in ws[rng]:
        for cell in row:
            cell.border = _box


def build_default_packing_list(po: dict, options: dict | None = None) -> bytes:
    """Generate a packing list xlsx matching the SSK template.

    `po` is the full PO document (dict) with `line_items`.
    `options` may carry `{carton_dim, pcs_per_box, net_wt_per_carton, gross_wt_per_carton}`.
    """
    options = options or {}
    pcs_per_box = int(options.get("pcs_per_box") or 20)
    net_wt = float(options.get("net_wt_per_carton") or 10.8)
    gross_wt = float(options.get("gross_wt_per_carton") or 12.0)
    carton_dim = options.get("carton_dim") or "60x50x30 CMS"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing list"

    # Determine the set of unique sizes across all line-items (sorted)
    sizes: list[str] = []
    for li in po.get("line_items", []):
        sz = str(li.get("size", "")).strip()
        if sz and sz not in sizes:
            sizes.append(sz)
    sizes.sort(key=lambda s: (len(s), s))
    if not sizes:
        sizes = ["—"]

    # ---- Title (A1:Q1) ----
    ws.merge_cells("A1:Q1")
    _set(ws, "A1", "PACKING LIST", bold=True, size=18, align="center", fill=_DARK, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    # ---- Vendor block A2:F6 (label A column, values B onwards) ----
    _set(ws, "A2", "VENDOR NAME :", bold=True, size=9, color=_ACCENT)
    ws.merge_cells("B2:F2"); _set(ws, "B2", VENDOR["name"], bold=True, size=11)
    ws.merge_cells("B3:F3"); _set(ws, "B3", VENDOR["address"], size=9)
    ws.merge_cells("B4:F4")
    _set(ws, "A6", "GSTIN:", bold=True, size=9, color=_ACCENT)
    ws.merge_cells("B6:F6"); _set(ws, "B6", VENDOR["gstin"], size=10, bold=True)

    # ---- Destination block G2:Q6 ----
    ws.merge_cells("G2:Q2")
    _set(ws, "G2", "DESTINATION HUB", bold=True, size=10, color=_ACCENT, align="left")
    ws.merge_cells("G3:Q3"); _set(ws, "G3", po.get("client_name", ""), bold=True, size=11)
    ws.merge_cells("G4:Q4"); _set(ws, "G4", po.get("client_address") or po.get("shipping_address", ""), size=9)
    ws.merge_cells("G5:Q5"); _set(ws, "G5", "", size=9)
    ws.merge_cells("G6:Q6"); _set(ws, "G6", f"GSTIN:- {po.get('client_gstin', '')}", size=10)

    # ---- Configuration row (PO No / Date / Carton Dim / box count) ----
    ws.merge_cells("A13:B13")
    _set(ws, "A13", "PO NO", bold=True, size=10, fill=_LIGHT, color=_ACCENT, align="left", border=True)
    ws.merge_cells("C13:E13")
    _set(ws, "C13", po.get("po_number", ""), bold=True, size=11, border=True)

    total_qty = int(po.get("total_quantity") or sum(int(li.get("quantity") or 0) for li in po.get("line_items", [])))
    total_cartons = max(1, (total_qty + pcs_per_box - 1) // pcs_per_box)

    _set(ws, "F13", total_qty, bold=True, size=10, border=True, align="center")
    _set(ws, "G13", "PCS", bold=True, size=10, border=True, align="center", color=_ACCENT)
    _set(ws, "H13", "BOX", bold=True, size=10, border=True, align="center", color=_ACCENT)
    _set(ws, "I13", total_cartons, bold=True, size=10, border=True, align="center")

    ws.merge_cells("A14:B14")
    _set(ws, "A14", "PO DATE", bold=True, size=10, fill=_LIGHT, color=_ACCENT, border=True)
    ws.merge_cells("C14:E14")
    _set(ws, "C14", po.get("po_date", ""), bold=True, size=11, border=True)
    ws.merge_cells("F14:K14")
    _set(ws, "F14", "", border=True)
    ws.merge_cells("L14:O14")
    _set(ws, "L14", "CARTON DIMENSION", bold=True, size=10, fill=_LIGHT, color=_ACCENT, align="center", border=True)
    ws.merge_cells("P14:Q14")
    _set(ws, "P14", carton_dim, bold=True, size=10, border=True, align="center")

    # ---- Line-item header row (row 16) ----
    headers = ["SITE CODE", "STYLE", "COLOUR", "CTN .NO"] + sizes + ["PCS/CTN", "PER CARTON", "TTL CTN", "TOTAL PCS", "NET WT", "GROSS WT"]
    # Make sure we extend to column Q max; if not enough cols, append blanks
    max_cols = max(len(headers), 17)
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        _set(ws, f"{col}16", h, bold=True, size=9, fill=_DARK, color="FFFFFF", align="center", border=True)
    ws.row_dimensions[16].height = 28

    # ---- Lines (rows 17+, one per non-zero PO line by (style,color)) ----
    # Aggregate quantities per (style_code, color) across sizes
    agg: dict[tuple[str, str], dict] = {}
    for li in po.get("line_items", []):
        key = (li.get("style_code", ""), li.get("color", ""))
        slot = agg.setdefault(key, {"style": key[0], "color": key[1], "by_size": {s: 0 for s in sizes}, "total": 0})
        sz = str(li.get("size", "")).strip()
        qty = int(li.get("quantity") or 0)
        if sz in slot["by_size"]:
            slot["by_size"][sz] += qty
        slot["total"] += qty

    n_size_cols = len(sizes)
    size_start_col_idx = 5  # column E
    pcs_ctn_col = size_start_col_idx + n_size_cols                                   # PCS/CTN
    per_ctn_col = pcs_ctn_col + 1                                                    # PER CARTON
    ttl_ctn_col = per_ctn_col + 1                                                    # TTL CTN
    total_pcs_col = ttl_ctn_col + 1                                                  # TOTAL PCS
    net_wt_col = total_pcs_col + 1                                                   # NET WT
    gross_wt_col = net_wt_col + 1                                                    # GROSS WT

    row_idx = 17
    ctn_seq = 1
    for (_, _), rec in agg.items():
        cartons_here = max(1, (rec["total"] + pcs_per_box - 1) // pcs_per_box)
        _set(ws, f"A{row_idx}", po.get("site_code", "—"), size=9, align="center", border=True)
        _set(ws, f"B{row_idx}", rec["style"], size=9, bold=True, align="left", border=True)
        _set(ws, f"C{row_idx}", rec["color"], size=9, align="left", border=True)
        _set(ws, f"D{row_idx}", str(ctn_seq), size=9, align="center", border=True)
        for i, sz in enumerate(sizes):
            col = get_column_letter(size_start_col_idx + i)
            _set(ws, f"{col}{row_idx}", rec["by_size"].get(sz, 0) or "", size=9, align="center", border=True)
        _set(ws, f"{get_column_letter(pcs_ctn_col)}{row_idx}", rec["total"], size=9, align="center", bold=True, border=True)
        _set(ws, f"{get_column_letter(per_ctn_col)}{row_idx}", pcs_per_box, size=9, align="center", border=True)
        _set(ws, f"{get_column_letter(ttl_ctn_col)}{row_idx}", cartons_here, size=9, align="center", border=True)
        _set(ws, f"{get_column_letter(total_pcs_col)}{row_idx}", rec["total"], size=9, align="center", bold=True, border=True)
        _set(ws, f"{get_column_letter(net_wt_col)}{row_idx}", net_wt, size=9, align="center", border=True)
        _set(ws, f"{get_column_letter(gross_wt_col)}{row_idx}", gross_wt, size=9, align="center", border=True)
        row_idx += 1
        ctn_seq += cartons_here

    # ---- Grand total row ----
    gt_row = row_idx
    _set(ws, f"A{gt_row}", "", border=True)
    _set(ws, f"B{gt_row}", "GRAND TOTAL", bold=True, size=10, fill=_DARK, color="FFFFFF", align="right", border=True)
    _set(ws, f"C{gt_row}", "", fill=_DARK, border=True)
    _set(ws, f"D{gt_row}", "", fill=_DARK, border=True)
    for i in range(n_size_cols):
        col = get_column_letter(size_start_col_idx + i)
        col_letter = get_column_letter(size_start_col_idx + i)
        _set(ws, f"{col_letter}{gt_row}", f"=SUM({col_letter}17:{col_letter}{gt_row-1})", bold=True, size=10, fill=_LIGHT, align="center", border=True)
    _set(ws, f"{get_column_letter(pcs_ctn_col)}{gt_row}", f"=SUM({get_column_letter(pcs_ctn_col)}17:{get_column_letter(pcs_ctn_col)}{gt_row-1})", bold=True, size=10, fill=_LIGHT, align="center", border=True)
    _set(ws, f"{get_column_letter(per_ctn_col)}{gt_row}", "", border=True, fill=_LIGHT)
    _set(ws, f"{get_column_letter(ttl_ctn_col)}{gt_row}", f"=SUM({get_column_letter(ttl_ctn_col)}17:{get_column_letter(ttl_ctn_col)}{gt_row-1})", bold=True, size=10, fill=_LIGHT, align="center", border=True)
    _set(ws, f"{get_column_letter(total_pcs_col)}{gt_row}", f"=SUM({get_column_letter(total_pcs_col)}17:{get_column_letter(total_pcs_col)}{gt_row-1})", bold=True, size=11, fill=_ACCENT, color="FFFFFF", align="center", border=True)
    _set(ws, f"{get_column_letter(net_wt_col)}{gt_row}", f"=SUM({get_column_letter(net_wt_col)}17:{get_column_letter(net_wt_col)}{gt_row-1})", bold=True, size=10, fill=_LIGHT, align="center", border=True)
    _set(ws, f"{get_column_letter(gross_wt_col)}{gt_row}", f"=SUM({get_column_letter(gross_wt_col)}17:{get_column_letter(gross_wt_col)}{gt_row-1})", bold=True, size=10, fill=_LIGHT, align="center", border=True)

    # ---- Order summary recap ----
    s_row = gt_row + 4
    ws.merge_cells(f"B{s_row}:B{s_row+3}")
    _set(ws, f"B{s_row}", "ORDER SUMMARY", bold=True, size=12, fill=_DARK, color="FFFFFF", align="center", border=True)
    _set(ws, f"D{s_row}", "Size", bold=True, size=10, fill=_LIGHT, color=_ACCENT, align="center", border=True)
    for i, sz in enumerate(sizes):
        col = get_column_letter(size_start_col_idx + i)
        _set(ws, f"{col}{s_row}", sz, bold=True, size=10, fill=_LIGHT, align="center", border=True)
    _set(ws, f"{get_column_letter(size_start_col_idx + n_size_cols)}{s_row}", "TOTAL", bold=True, size=10, fill=_DARK, color="FFFFFF", align="center", border=True)

    for label_row, label in enumerate(["Order Qty", "Pack Qty", "Excess / Short"], start=1):
        r = s_row + label_row
        _set(ws, f"D{r}", label, bold=True, size=9, align="right", border=True)
        for i in range(n_size_cols):
            col = get_column_letter(size_start_col_idx + i)
            if label == "Order Qty":
                val = f"={col}{gt_row}"
            elif label == "Pack Qty":
                val = f"={col}{gt_row}"
            else:
                val = f"={col}{s_row+2}-{col}{s_row+1}"
            _set(ws, f"{col}{r}", val, size=9, align="center", border=True)
        tot_col = get_column_letter(size_start_col_idx + n_size_cols)
        _set(ws, f"{tot_col}{r}", f"=SUM({get_column_letter(size_start_col_idx)}{r}:{get_column_letter(size_start_col_idx + n_size_cols - 1)}{r})", bold=True, size=9, fill=_LIGHT, align="center", border=True)

    # Column widths
    for col_idx in range(1, max_cols + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 12 if col_idx in (1, 2, 3) else 9

    # Save & return bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Custom template substitution ----------------------------------------------
# ---------------------------------------------------------------------------
_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z_]+)\s*\}\}")


def build_from_template(template_bytes: bytes, po: dict, options: dict | None = None) -> bytes:
    """Fill an uploaded template file. The template must contain placeholders
    such as ``{{po_number}}``. The special placeholder ``{{lines}}`` (alone in
    a cell) marks the first line-item row; that row gets cloned downwards for
    each line item and the immediately-preceding row is treated as its header,
    used to map column positions to fields.
    """
    options = options or {}
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    today = datetime.now().strftime("%d.%m.%Y")
    total_qty = int(po.get("total_quantity") or sum(int(li.get("quantity") or 0) for li in po.get("line_items", [])))
    pcs_per_box = int(options.get("pcs_per_box") or 20)
    total_cartons = max(1, (total_qty + pcs_per_box - 1) // pcs_per_box)

    scalars = {
        "po_number": po.get("po_number", ""),
        "po_date": po.get("po_date", ""),
        "client_name": po.get("client_name", ""),
        "client_address": po.get("client_address", "") or po.get("shipping_address", ""),
        "client_gstin": po.get("client_gstin", ""),
        "vendor_name": VENDOR["name"],
        "vendor_address": VENDOR["address"],
        "vendor_gstin": VENDOR["gstin"],
        "carton_dim": options.get("carton_dim", "60x50x30 CMS"),
        "date": today,
        "total_pcs": total_qty,
        "total_cartons": total_cartons,
    }

    for ws in wb.worksheets:
        _substitute_scalars(ws, scalars)
        _expand_lines(ws, po, options)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _substitute_scalars(ws, scalars: dict) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            new_val = _PLACEHOLDER_RE.sub(lambda m: str(scalars.get(m.group(1), m.group(0))), cell.value)
            if new_val != cell.value:
                cell.value = new_val


def _find_lines_marker(ws) -> Optional[tuple[int, int]]:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{lines}}" in cell.value:
                return cell.row, cell.column
    return None


def _classify_header_simple(s: str) -> Optional[str]:
    if not s:
        return None
    s = re.sub(r"\s+", " ", str(s).strip().lower())
    table = {
        "style": ["style", "article", "model"],
        "description": ["description", "particulars", "product"],
        "color": ["color", "colour"],
        "size": ["size"],
        "quantity": ["qty", "quantity", "pcs", "pairs", "total pcs"],
        "unit_price": ["rate", "price", "unit price"],
        "amount": ["amount", "total"],
        "net_wt": ["net wt", "net weight"],
        "gross_wt": ["gross wt", "gross weight"],
        "ctn_no": ["ctn no", "carton no", "ctn.no", "ctn .no", "carton number"],
    }
    for key, choices in table.items():
        for c in choices:
            if c == s or (len(s) <= 20 and c in s):
                return key
    return None


def _expand_lines(ws, po: dict, options: dict | None = None) -> None:
    options = options or {}
    pcs_per_box = int(options.get("pcs_per_box") or 20)
    marker = _find_lines_marker(ws)
    if marker is None:
        return
    marker_row, _ = marker
    # The header is the previous non-empty row above the marker
    header_row = marker_row - 1
    while header_row > 0:
        vals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in vals):
            break
        header_row -= 1
    header_map: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        cls = _classify_header_simple(ws.cell(header_row, c).value)
        if cls:
            header_map[c] = cls

    # Aggregate PO line items per (style, color) and pull line records
    items = po.get("line_items", [])
    if not items:
        return

    # Wipe the marker row first
    template_row_height = ws.row_dimensions[marker_row].height
    for c in range(1, ws.max_column + 1):
        ws.cell(marker_row, c).value = None

    # Write each line item into a row, starting from marker_row
    for offset, li in enumerate(items):
        target_row = marker_row + offset
        if offset > 0:
            ws.insert_rows(target_row)
            ws.row_dimensions[target_row].height = template_row_height
            # Copy styles from marker_row to target_row
            for c in range(1, ws.max_column + 1):
                src = ws.cell(marker_row, c)
                dst = ws.cell(target_row, c)
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.alignment = copy(src.alignment)
                    dst.border = copy(src.border)
                    dst.number_format = src.number_format

        qty = int(li.get("quantity") or 0)
        cartons = max(1, (qty + pcs_per_box - 1) // pcs_per_box)
        field_values = {
            "style": li.get("style_code", ""),
            "description": li.get("description", ""),
            "color": li.get("color", ""),
            "size": li.get("size", ""),
            "quantity": qty,
            "unit_price": li.get("unit_price", 0),
            "amount": li.get("amount", round(qty * float(li.get("unit_price") or 0), 2)),
            "net_wt": float(options.get("net_wt_per_carton") or 10.8) * cartons,
            "gross_wt": float(options.get("gross_wt_per_carton") or 12.0) * cartons,
            "ctn_no": offset + 1,
        }
        for c, field in header_map.items():
            ws.cell(target_row, c).value = field_values.get(field)
