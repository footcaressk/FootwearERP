"""AI-powered PO extraction from PDF / Excel files using Emergent LLM key + Gemini."""
import os
import json
import re
import tempfile
from pathlib import Path
import openpyxl
from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType


EXTRACTION_PROMPT = """You are an expert at extracting structured data from Purchase Order documents for footwear manufacturing.

Extract the following fields from the attached PO document and return ONLY valid JSON (no markdown, no commentary):

{
  "po_number": "string",
  "po_date": "YYYY-MM-DD",
  "client_name": "string",
  "client_address": "string",
  "vendor_name": "string",
  "vendor_address": "string",
  "billing_address": "string",
  "shipping_address": "string",
  "delivery_date": "YYYY-MM-DD",
  "payment_terms": "string",
  "currency": "INR",
  "line_items": [
    {
      "item_code": "string",
      "style_code": "string",
      "description": "string",
      "color": "string",
      "size": "string",
      "hsn_code": "string",
      "quantity": 0,
      "unit_price": 0.0,
      "amount": 0.0
    }
  ],
  "subtotal": 0.0,
  "cgst_rate": 0.0,
  "cgst_amount": 0.0,
  "sgst_rate": 0.0,
  "sgst_amount": 0.0,
  "igst_rate": 0.0,
  "igst_amount": 0.0,
  "total_tax": 0.0,
  "grand_total": 0.0,
  "total_quantity": 0,
  "notes": "string"
}

Rules:
- Convert dates from DD.MM.YYYY or DD/MM/YYYY to YYYY-MM-DD.
- For interstate transactions use IGST, intrastate use CGST+SGST.
- Each size variant should be its own line_item.
- All numeric values must be numbers, not strings.
- If a field is missing, use null or empty string.
- Return ONLY the JSON, no other text."""


def xlsx_to_text(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"### Sheet: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                parts.append(" | ".join(cells))
    return "\n".join(parts)


def _clean_json(text: str) -> str:
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return text.strip()


async def extract_po_from_pdf(file_bytes: bytes) -> dict:
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise RuntimeError("EMERGENT_LLM_KEY not configured")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"po-extract-{os.urandom(8).hex()}",
            system_message="You extract structured data from purchase orders and return strict JSON.",
        ).with_model("gemini", "gemini-2.5-flash")

        file_attach = FileContentWithMimeType(file_path=tmp_path, mime_type="application/pdf")
        response = await chat.send_message(
            UserMessage(text=EXTRACTION_PROMPT, file_contents=[file_attach])
        )
        cleaned = _clean_json(response)
        return json.loads(cleaned)
    finally:
        Path(tmp_path).unlink(missing_ok=True)


async def extract_po_from_xlsx(file_bytes: bytes) -> dict:
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise RuntimeError("EMERGENT_LLM_KEY not configured")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        text_content = xlsx_to_text(tmp_path)
        chat = LlmChat(
            api_key=api_key,
            session_id=f"po-extract-{os.urandom(8).hex()}",
            system_message="You extract structured data from purchase orders and return strict JSON.",
        ).with_model("gemini", "gemini-2.5-flash")

        response = await chat.send_message(
            UserMessage(text=f"{EXTRACTION_PROMPT}\n\n--- Document Content ---\n{text_content}")
        )
        cleaned = _clean_json(response)
        return json.loads(cleaned)
    finally:
        Path(tmp_path).unlink(missing_ok=True)
